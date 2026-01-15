import sys
import io

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Use openpyxl to read the Excel-XML file
from openpyxl import load_workbook

file_path = r'C:\Users\irfan\Downloads\weekly_stocks_all_2025-12-08.xls'

print("=" * 100)
print("ANALYZING EXCEL FILE: weekly_stocks_all_2025-12-08.xls")
print("=" * 100)

try:
    # Try to open as Excel XML
    wb = load_workbook(file_path, read_only=True, data_only=True)

    print(f"\nSheet names found: {wb.sheetnames}\n")

    # Analyze Structure sheet
    if 'Structure' in wb.sheetnames:
        ws_structure = wb['Structure']
        print("\n" + "=" * 100)
        print("STRUCTURE TAB (Schema Definition)")
        print("=" * 100)
        print(f"Total rows in Structure: {ws_structure.max_row}")
        print(f"Total columns in Structure: {ws_structure.max_column}\n")

        print("Field definitions (in order as defined in Structure tab):")
        print("-" * 100)
        structure_fields = []
        for i, row in enumerate(ws_structure.iter_rows(min_row=2, max_row=min(100, ws_structure.max_row), values_only=True), 1):
            if row[1]:  # If Field Name exists
                field_num = row[0]
                field_name = row[1]
                field_type = row[2]
                field_mode = row[3]
                structure_fields.append(field_name)
                print(f"{field_num:3}. {field_name:40} | Type: {field_type:10} | Mode: {field_mode}")

    # Analyze Data sheet
    if 'Data' in wb.sheetnames:
        ws_data = wb['Data']
        print("\n" + "=" * 100)
        print("DATA TAB (Actual Data)")
        print("=" * 100)
        print(f"Total rows in Data: {ws_data.max_row}")
        print(f"Total columns in Data: {ws_data.max_column}\n")

        # Get header row
        data_fields = []
        header_row = next(ws_data.iter_rows(min_row=1, max_row=1, values_only=True))

        print("Column order in Data tab:")
        print("-" * 100)
        for i, col_name in enumerate(header_row, 1):
            if col_name:
                data_fields.append(col_name)
                print(f"{i:3}. {col_name}")

        # Compare Structure vs Data order
        print("\n" + "=" * 100)
        print("COMPARISON: Structure Tab vs Data Tab")
        print("=" * 100)

        print(f"\nTotal fields in Structure: {len(structure_fields)}")
        print(f"Total fields in Data:      {len(data_fields)}")

        if structure_fields == data_fields:
            print("\n✅ MATCH: Field order is IDENTICAL in both tabs")
        else:
            print("\n❌ MISMATCH: Field order is DIFFERENT between tabs\n")

            # Find differences
            print("Detailed comparison:")
            print("-" * 100)
            max_len = max(len(structure_fields), len(data_fields))

            for i in range(max_len):
                struct_field = structure_fields[i] if i < len(structure_fields) else "MISSING"
                data_field = data_fields[i] if i < len(data_fields) else "MISSING"

                match = "✓" if struct_field == data_field else "✗"

                print(f"{i+1:3}. {match} | Structure: {struct_field:40} | Data: {data_field}")

            # Fields in Structure but not in Data
            missing_in_data = set(structure_fields) - set(data_fields)
            if missing_in_data:
                print(f"\n⚠️ Fields in Structure but NOT in Data: {missing_in_data}")

            # Fields in Data but not in Structure
            extra_in_data = set(data_fields) - set(structure_fields)
            if extra_in_data:
                print(f"\n⚠️ Fields in Data but NOT in Structure: {extra_in_data}")

        # Show sample data
        print("\n" + "=" * 100)
        print("SAMPLE DATA (First 3 rows)")
        print("=" * 100)
        for i, row in enumerate(ws_data.iter_rows(min_row=2, max_row=4, values_only=True), 1):
            print(f"\nRow {i}:")
            for j, (col_name, value) in enumerate(zip(data_fields, row)):
                if j < 10:  # Show first 10 columns only
                    print(f"  {col_name}: {value}")

    wb.close()

except Exception as e:
    print(f"\n❌ Error reading file: {e}")
    print(f"Error type: {type(e).__name__}")

    # Try alternative method with pandas
    print("\nTrying alternative method with pandas...")
    import pandas as pd

    try:
        # Read with pandas
        df_structure = pd.read_excel(file_path, sheet_name='Structure', engine=None)
        df_data = pd.read_excel(file_path, sheet_name='Data', nrows=5, engine=None)

        print("\nStructure columns:", list(df_structure.columns))
        print("\nData columns:", list(df_data.columns))

    except Exception as e2:
        print(f"Pandas also failed: {e2}")
